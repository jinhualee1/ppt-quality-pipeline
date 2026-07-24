from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from .io import read_json, write_json


def _annotation_summary(run_dir: Path) -> dict[str, str]:
    path = run_dir / "annotations.json"
    if not path.is_file():
        return {}
    values = read_json(path)
    summary: dict[str, list[str]] = {}
    for key, value in values.items():
        labels = value.get("labels", [])
        note = value.get("note", "").strip()
        description = ", ".join(labels)
        if note:
            description = f"{description}: {note}" if description else note
        item_id = key.split("/", 1)[0]
        if description:
            summary.setdefault(item_id, []).append(f"{key}: {description}")
    return {key: "; ".join(items) for key, items in summary.items()}


def report_rows(run_dir: Path) -> list[dict[str, Any]]:
    report = read_json(run_dir / "report.json")
    annotations = _annotation_summary(run_dir)
    rows = []
    for item in report["items"]:
        issues = "; ".join(issue["message"] for issue in item["issues"]) or "No automated issues"
        pages = sum(deck["page_count"] for deck in item["decks"] if deck["status"] == "rendered")
        renderers = ", ".join(dict.fromkeys(deck["renderer"] for deck in item["decks"] if deck["renderer"]))
        fidelity = ", ".join(dict.fromkeys(deck["fidelity"] for deck in item["decks"] if deck["fidelity"]))
        rows.append(
            {
                "id": item["id"],
                "status": item["status"],
                "query": item["query"],
                "rendered_pages": pages,
                "renderers": renderers,
                "fidelity": fidelity,
                "automated_issues": issues,
                "manual_annotations": annotations.get(item["id"], ""),
            }
        )
    return rows


def export_csv(run_dir: Path, output: Path | None = None) -> Path:
    output = output or run_dir / "report.csv"
    rows = report_rows(run_dir)
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "id",
        "status",
        "query",
        "rendered_pages",
        "renderers",
        "fidelity",
        "automated_issues",
        "manual_annotations",
    ]
    with output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return output


def export_xlsx(run_dir: Path, output: Path | None = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("XLSX export requires: pip install -e .[xlsx]") from exc

    output = output or run_dir / "report.xlsx"
    rows = report_rows(run_dir)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PPT Quality Report"
    headers = [
        "ID",
        "Status",
        "Query",
        "Rendered pages",
        "Renderers",
        "Fidelity",
        "Automated issues",
        "Manual annotations",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17212B")
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append(
            [
                row["id"],
                row["status"],
                row["query"],
                row["rendered_pages"],
                row["renderers"],
                row["fidelity"],
                row["automated_issues"],
                row["manual_annotations"],
            ]
        )
    widths = [18, 14, 58, 18, 28, 14, 72, 72]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def export_json_summary(run_dir: Path, output: Path | None = None) -> Path:
    output = output or run_dir / "report.summary.json"
    write_json(output, {"items": report_rows(run_dir)})
    return output
